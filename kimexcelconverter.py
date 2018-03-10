import openpyxl

def run():
    print("starting...")
    data = UseOpenpyxl("KimTestData.xlsx")
    newStudentList = PivotSecondParentData(data)
    TransformAndExport("KimExportedData.xlsx", newStudentList)
    print("done")

def PivotSecondParentData(dataList):
    studentHash = {}
    # Hash of all students to their parental contacts
    for dataRow in dataList:
        if not dataRow["ChildID"] in studentHash:
            studentHash[dataRow["ChildID"]] = []
        studentHash[dataRow["ChildID"]].append(dataRow)

    newStudentList = []
    for key, value in studentHash.items():
        firstContact = value[0]
        newStudentList.append(firstContact)
        if (len(value) > 1):
            secondContact = value[1]
            firstContact["Parent2FirstName"] = secondContact["ParentFirstName"]
            firstContact["Parent2LastName"] = secondContact["ParentLastName"]
            firstContact["Parent2Language"] = secondContact["ParentLanguage"]
            firstContact["Parent2Gender"] = secondContact["ParentGender"]
            firstContact["Family2Phone"] = secondContact["FamilyPhone"]
            firstContact["Family2Address1"] = secondContact["FamilyAddress1"]
            firstContact["Family2City"] = secondContact["FamilyCity"]
            firstContact["Family2Zip"] = secondContact["FamilyZip"]
            firstContact["Parent2Relationship"] = secondContact["ParentRelationship"]
    return newStudentList

def TransformAndExport(dest_file_name, dataList):
    wb = openpyxl.workbook.Workbook()
    sheet = wb.active
    WriteHeader(sheet)
    rowNum = 2
    for dataRow in dataList:
        WriteRows(sheet, rowNum, dataRow)
        rowNum+=1
    wb.save(filename=dest_file_name)



def WriteRows(sheet, rowNum, dataRow):
    sheet.cell(row=rowNum, column=1).value = dataRow["ChildID"] #"Student ID"
    sheet.cell(row=rowNum, column=2).value = "" #"State Identifier"
    sheet.cell(row=rowNum, column=3).value = dataRow["ChildFirstName"] #"Student's First Name"
    sheet.cell(row=rowNum, column=4).value = dataRow["ChildLastName"] #"Student's Last Name"
    sheet.cell(row=rowNum, column=5).value = dataRow["ChildGender"] #"Student's Gender"
    sheet.cell(row=rowNum, column=6).value = dataRow["ChildDateofBirth"] #"Student's DOB"
    sheet.cell(row=rowNum, column=7).value = "" #"Ethnicity"
    sheet.cell(row=rowNum, column=8).value = "" #"Student's Age"
    sheet.cell(row=rowNum, column=9).value = "" #"Grade Level"
    sheet.cell(row=rowNum, column=10).value = "" #"Class/Cohert"
    sheet.cell(row=rowNum, column=11).value = "" #"Foster"
    sheet.cell(row=rowNum, column=12).value = "" #"McKinney-Vento/Homeless"
    sheet.cell(row=rowNum, column=13).value = "" #"Migrant Ed"
    sheet.cell(row=rowNum, column=14).value = "" #"Special Ed"
    sheet.cell(row=rowNum, column=15).value = dataRow["CenterName"] #"School Name"
    sheet.cell(row=rowNum, column=16).value = "" #"Teacher"
    sheet.cell(row=rowNum, column=17).value = dataRow["ParentFirstName"] #"Parent's First Name"
    sheet.cell(row=rowNum, column=18).value = dataRow["ParentLastName"] #"Parent's Last Name"
    sheet.cell(row=rowNum, column=19).value = dataRow["ParentLanguage"] #"Parent's Preferred Language"
    sheet.cell(row=rowNum, column=20).value = dataRow["ParentGender"] #"Parent's Gender"
    sheet.cell(row=rowNum, column=21).value = dataRow["ParentRelationship"]  # "Relationship to Child"
    sheet.cell(row=rowNum, column=22).value = dataRow["FamilyPhone"] #"Parent Main Phone"
    sheet.cell(row=rowNum, column=23).value = "" #"Parent's Home Phone"
    sheet.cell(row=rowNum, column=24).value = dataRow["FamilyAddress1"] #"Street/Mailing Address"
    sheet.cell(row=rowNum, column=25).value = dataRow["FamilyCity"] #"Mailing City"
    sheet.cell(row=rowNum, column=26).value = "" #"Mailing State/Province"
    sheet.cell(row=rowNum, column=27).value = dataRow["FamilyZip"] #"Mailing Zip/Postal Code"

    sheet.cell(row=rowNum, column=28).value = dataRow["Parent2FirstName"] #"Second Parent's First Name"
    sheet.cell(row=rowNum, column=29).value = dataRow["Parent2LastName"] #"Second Parent's Last Name"
    sheet.cell(row=rowNum, column=30).value = dataRow["Parent2Gender"] #"Second Parent's Gender"
    sheet.cell(row=rowNum, column=31).value = dataRow["Parent2Relationship"] #"Second Parent's Relationship to Child"
    sheet.cell(row=rowNum, column=32).value = dataRow["Parent2Language"] #"Parent's Preferred Language"
    sheet.cell(row=rowNum, column=33).value = dataRow["Family2Phone"] #"Second Parent's Phone"
    sheet.cell(row=rowNum, column=34).value = dataRow["Family2Address1"] #"Street/Mailing Address"
    sheet.cell(row=rowNum, column=35).value = dataRow["Family2City"] #"Mailing City"
    sheet.cell(row=rowNum, column=36).value = "" #"Mailing State/Province"
    sheet.cell(row=rowNum, column=37).value = dataRow["Family2Zip"] #"Mailing Zip/Postal Code"

def WriteHeader(sheet):
    sheet.cell(row=1, column=1).value = "Student ID"
    sheet.cell(row=1, column=2).value = "State Identifier"
    sheet.cell(row=1, column=3).value = "Student's First Name"
    sheet.cell(row=1, column=4).value = "Student's Last Name"
    sheet.cell(row=1, column=5).value = "Student's Gender"
    sheet.cell(row=1, column=6).value = "Student's DOB"
    sheet.cell(row=1, column=7).value = "Ethnicity"
    sheet.cell(row=1, column=8).value = "Student's Age"
    sheet.cell(row=1, column=9).value = "Grade Level"
    sheet.cell(row=1, column=10).value = "Class/Cohert"
    sheet.cell(row=1, column=11).value = "Foster"
    sheet.cell(row=1, column=12).value = "McKinney-Vento/Homeless"
    sheet.cell(row=1, column=13).value = "Migrant Ed"
    sheet.cell(row=1, column=14).value = "Special Ed"
    sheet.cell(row=1, column=15).value = "School Name"
    sheet.cell(row=1, column=16).value = "Teacher"
    sheet.cell(row=1, column=17).value = "Parent's First Name"
    sheet.cell(row=1, column=18).value = "Parent's Last Name"
    sheet.cell(row=1, column=19).value = "Parent's Preferred Language"
    sheet.cell(row=1, column=20).value = "Parent's Gender"
    sheet.cell(row=1, column=21).value = "Relationship to Child"
    sheet.cell(row=1, column=22).value = "Parent Main Phone"
    sheet.cell(row=1, column=23).value = "Parent's Home Phone"
    sheet.cell(row=1, column=24).value = "Street/Mailing Address"
    sheet.cell(row=1, column=25).value = "Mailing City"
    sheet.cell(row=1, column=26).value = "Mailing State/Province"
    sheet.cell(row=1, column=27).value = "Mailing Zip/Postal Code"
    sheet.cell(row=1, column=28).value = "Second Parent's First Name"
    sheet.cell(row=1, column=29).value = "Second Parent's Last Name"
    sheet.cell(row=1, column=30).value = "Second Parent's Gender"
    sheet.cell(row=1, column=31).value = "Second Parent's Relationship to Child"
    sheet.cell(row=1, column=32).value = "Parent's Preferred Language"
    sheet.cell(row=1, column=33).value = "Second Parent's Phone"
    sheet.cell(row=1, column=34).value = "Street/Mailing Address"
    sheet.cell(row=1, column=35).value = "Mailing City"
    sheet.cell(row=1, column=36).value = "Mailing State/Province"
    sheet.cell(row=1, column=37).value = "Mailing Zip/Postal Code"


def UseOpenpyxl(file_name):
    wb = openpyxl.load_workbook(file_name, read_only=True)
    sheet = wb.active
    rows = sheet.rows
    first_row = [cell.value for cell in next(rows)]
    data = []
    for row in rows:
        record = {}
        for key, cell in zip(first_row, row):
            if cell.data_type == 's':
                record[key] = cell.value.strip()
            else:
                record[key] = cell.value
        data.append(record)
    return data


run()
